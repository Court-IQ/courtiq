#!/usr/bin/env python3
"""Build programs.json from programs.xlsx (v3 schema).

One row per drill, grouped by workout_id. All sheets are merged into one list.

Usage:
    cd ~/courtiq && python3 programs/build_programs.py
"""

from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
SHEET = ROOT / "programs.xlsx"
OUT = ROOT.parent / "public" / "programs.json"

REQUIRED = {
    "workout_id",
    "workout_title",
    "byline",
    "level",
    "focuses",
    "minutes_per_day",
    "summary",
    "drill_order",
    "drill_name",
    "drill_reps",
    "drill_description",
}


def split_list(value) -> list[str]:
    if value is None:
        return []
    return [v.strip() for v in str(value).split("/") if v.strip()]


def cell_str(v) -> str:
    return "" if v is None else str(v).strip()


def main():
    if not SHEET.exists():
        print(f"ERROR: {SHEET} not found.")
        raise SystemExit(1)

    wb = openpyxl.load_workbook(SHEET, data_only=True)
    workouts: "OrderedDict[str, dict]" = OrderedDict()
    errors: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            continue
        header_cells = rows[0]
        headers = [cell_str(c.value) for c in header_cells]
        missing = REQUIRED - set(headers)
        if missing:
            errors.append(f"Sheet '{sheet_name}' missing columns: {missing}")
            continue
        col = {h: i for i, h in enumerate(headers)}

        for row_idx, row in enumerate(rows[1:], start=2):
            values = [c.value for c in row]
            workout_id = cell_str(values[col["workout_id"]])
            if not workout_id:
                continue

            drill_order = values[col["drill_order"]]
            try:
                drill_order_int = int(drill_order)
            except (TypeError, ValueError):
                errors.append(
                    f"{sheet_name} row {row_idx}: drill_order must be a number"
                )
                continue

            if workout_id not in workouts:
                try:
                    minutes_val = int(values[col["minutes_per_day"]])
                except (TypeError, ValueError):
                    errors.append(
                        f"{sheet_name} row {row_idx}: minutes_per_day must be a number"
                    )
                    continue

                workouts[workout_id] = {
                    "id": workout_id,
                    "title": cell_str(values[col["workout_title"]]),
                    "byline": cell_str(values[col["byline"]]) or "Coach Matthew",
                    "level": split_list(values[col["level"]]),
                    "focus": split_list(values[col["focuses"]]),
                    "minutes_per_day": minutes_val,
                    "summary": cell_str(values[col["summary"]]),
                    "days": [
                        {
                            "day_number": 1,
                            "title": "Today's Workout",
                            "total_minutes": minutes_val,
                            "drills": [],
                        }
                    ],
                }

            video_idx = col.get("drill_video_url")
            focus_idx = col.get("drill_focus")
            drill = {
                "name": cell_str(values[col["drill_name"]]),
                "description": cell_str(values[col["drill_description"]]),
                "reps_or_time": cell_str(values[col["drill_reps"]]),
                "video_url": cell_str(values[video_idx]) if video_idx is not None else "",
                "focus": cell_str(values[focus_idx]) if focus_idx is not None else "",
            }
            if drill["video_url"] == "":
                drill["video_url"] = None
            workouts[workout_id]["days"][0]["drills"].append((drill_order_int, drill))

    if errors:
        print("Build errors:")
        for e in errors:
            print(f"  - {e}")
        raise SystemExit(1)

    output_programs = []
    for w in workouts.values():
        for day in w["days"]:
            day["drills"].sort(key=lambda x: x[0])
            day["drills"] = [d for _, d in day["drills"]]
        output_programs.append(w)

    payload = {"version": 3, "programs": output_programs}
    OUT.write_text(json.dumps(payload, separators=(",", ":")))
    print(
        f"OK  {len(output_programs)} workouts  "
        f"{sum(len(d['drills']) for w in output_programs for d in w['days'])} drills  "
        f"-> {OUT}"
    )


if __name__ == "__main__":
    main()
