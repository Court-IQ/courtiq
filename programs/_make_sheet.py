#!/usr/bin/env python3
"""
Build programs.xlsx with 5 sheets — one per position.

Each sheet has the same column structure. Programs primarily aimed at a
position go in that sheet. The `position` cell in the row is still the
source of truth — a row can list multiple positions for matching, the
sheet is just where you maintain it.

After running this you can edit the .xlsx in Excel/Numbers and ADD rows.
The build_programs.py script merges all 5 sheets into one JSON.

WARNING — running this script wipes the .xlsx and rebuilds from the
hardcoded example programs below. Only run it again when changing the
column structure or sheet layout. For ongoing edits, edit the .xlsx
directly.

Usage:
    python3 programs/_make_sheet.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "programs.xlsx"

# Sheet names — match position numbers + label for clarity
SHEETS = [
    ("1 - PG", "1"),
    ("2 - SG", "2"),
    ("3 - SF", "3"),
    ("4 - PF", "4"),
    ("5 - C",  "5"),
]

COLUMNS = [
    ("program_id",         18, "kebab-case unique id, e.g. pg-shoot-hs-45-5d"),
    ("program_title",      28, "Display title"),
    ("byline",             16, "'Coach Matthew' or attribution"),
    ("position",           20, "Comma-separated. Valid values: 1, 2, 3, 4, 5"),
    ("level",              20, "Comma-separated. Valid values: High School, College, Pro"),
    ("focus",              16, "Comma-separated. Valid values: Shooting, Ball handling, Finishing at the rim, Game IQ + film"),
    ("minutes_per_day",    14, "15, 30, 45, 60, or 90"),
    ("days_per_week",      14, "3-7"),
    ("summary",            40, "One-sentence summary of the program"),
    ("day_number",         12, "1-7"),
    ("day_title",          22, "Day's theme: 'Form + Volume', 'Game Speed', etc."),
    ("day_total_minutes",  16, "Should match minutes_per_day"),
    ("drill_order",        12, "1, 2, 3, 4 — order within the day"),
    ("drill_name",         32, "'Catch-and-shoot from 5 spots' — descriptive, not branded"),
    ("drill_reps",         24, "'50 makes' or '3 sets of 10' or '8 min'"),
    ("drill_description",  60, "1-3 sentence drill instructions with form cues"),
    ("drill_video_url",    40, "Optional YouTube URL demonstrating the drill. Leave blank if none."),
]

# ---------------------------------------------------------------------------
# Hardcoded example programs that live in sheet "1 - PG"
# ---------------------------------------------------------------------------
PG_SHOOTING = {
    "meta": {
        "program_id": "pg-shoot-hs-45-5d",
        "program_title": "Point Guard Shooting Series #1",
        "byline": "Coach Matthew",
        "position": "1, 2",
        "level": "High School",
        "focus": "Shooting",
        "minutes_per_day": 45,
        "days_per_week": 5,
        "summary": "Five days of game-speed shooting volume. Form work Day 1, off-dribble and game-speed reps as the week progresses.",
    },
    "days": [
        {
            "day_number": 1,
            "day_title": "Form + Volume",
            "day_total_minutes": 45,
            "drills": [
                ("Form shots from 6 feet", "50 makes (10 from each of 5 spots)",
                 "Stand 6 ft from rim. Shot pocket low, elbow under the ball, follow-through with full extension and hold. Hand under, not on the side."),
                ("Catch-and-shoot from 5 spots", "50 makes (10 per spot)",
                 "Spin pass to yourself, catch in shot pocket, square hips and shoulders to rim, shoot in rhythm. Don't dip the ball. Spots: left corner, left wing, top, right wing, right corner."),
                ("One-dribble pull-ups from each elbow", "20 makes (10 per side)",
                 "Start at the top of the key. One hard dribble to the elbow, gather, pull up. Footwork: inside foot plants first. Right side first, then left."),
            ],
        },
        {
            "day_number": 2,
            "day_title": "Off-Dribble Mid-Range",
            "day_total_minutes": 45,
            "drills": [
                ("Wing drive + pull-up", "20 makes (10 per side)",
                 "From the wing, hard dribble to the elbow, jab-style pull-up. Focus on shoulder square and balance. Right then left."),
                ("Hesitation into pull-up at the elbow", "20 makes (10 per side)",
                 "Slow-fast move — change pace mid-dribble, then pull. The hesitation sells it. Don't shortcut the gather."),
                ("Step-back from the wing", "20 makes (10 per side)",
                 "Drive baseline 2 steps, push off front foot, step back into shot pocket. Stay low through the move."),
            ],
        },
        {
            "day_number": 3,
            "day_title": "3-Point Volume + Recovery",
            "day_total_minutes": 45,
            "drills": [
                ("3-point catch-and-shoot from 5 spots", "100 makes total (20 per spot)",
                 "Same 5 spots as Day 1 but behind the line. Spin pass, catch ready, shoot. Strong base, no fade."),
                ("Off-the-dribble 3 from wing", "20 makes (10 per side)",
                 "One hard side-step dribble before the pull. Stay behind the line. Get used to the rhythm of one-dribble 3s."),
            ],
        },
        {
            "day_number": 4,
            "day_title": "Game Speed",
            "day_total_minutes": 45,
            "drills": [
                ("Spot-up shooting under closeout pressure", "50 reps total",
                 "Partner or chair closeout. Catch, rip-through, shoot or one-dribble pull depending on the closeout. Game decision."),
                ("Pull-up off pick simulation", "30 makes (15 per side)",
                 "Set imaginary screen at the top. Dribble off it, attack downhill, pull at the elbow. Get the angle right."),
                ("Free throws at the end (fatigued)", "25 makes",
                 "After all your shooting work, shoot FTs. Routine the same every time. Breathe. Bend knees same depth."),
            ],
        },
        {
            "day_number": 5,
            "day_title": "Make It Take It",
            "day_total_minutes": 45,
            "drills": [
                ("5-spot make-it streak", "Make 5 in a row from each of 5 spots",
                 "From each spot, make 5 in a row before moving. If you miss, start that spot over. Corner → wing → top → wing → corner. Mental rep as much as physical."),
                ("Free throws", "50 makes",
                 "Routine every time. Same dribbles, same breath, same release. Track the streak."),
                ("End-of-game shot simulation", "20 reps",
                 "5 seconds on imaginary clock. Count down out loud (3-2-1), catch-and-shoot a 3. Reset, repeat. Mental rep matters more than form here."),
            ],
        },
    ],
}

PG_HANDLE = {
    "meta": {
        "program_id": "pg-handle-hs-15-3d",
        "program_title": "Point Guard Ball Handling Series #1",
        "byline": "Coach Matthew",
        "position": "1",
        "level": "High School",
        "focus": "Ball handling",
        "minutes_per_day": 15,
        "days_per_week": 3,
        "summary": "Three short days for HS point guards — tight handles, weak hand confidence, and change-of-pace moves.",
    },
    "days": [
        {
            "day_number": 1,
            "day_title": "Stationary Foundations",
            "day_total_minutes": 15,
            "drills": [
                ("Pound dribble progression",
                 "3 rounds: 30s right / 30s left / 30s alternating",
                 "Pound the ball hard at hip height. Knees bent, eyes up, never look down at the ball. Right hand 30 sec, then left, then alternate every dribble for 30 sec. Rest 30 sec between rounds."),
                ("Stationary handle sequence",
                 "4 sets — 8 reps of each move",
                 "One set = crossover (8 reps), between the legs (8 reps), behind the back (8 reps), in-and-out (8 reps). Keep the dribble low and hard. Don't get sloppy when you're tired."),
                ("Weak hand pound — height changes",
                 "3 sets of 60 sec",
                 "Left hand only. Pound dribble and change heights every 15 sec — high, low, high, low. Eyes up the whole time. This is the drill that closes the gap between your hands."),
            ],
        },
        {
            "day_number": 2,
            "day_title": "On-the-Move",
            "day_total_minutes": 15,
            "drills": [
                ("Full-court speed dribble",
                 "10 trips — alternate hand each trip",
                 "Push dribble at top speed end-to-end. Stay low, push the ball out in front of you (not next to you). Trip 1 right, trip 2 left, repeat. Touch each baseline."),
                ("Hesitation move attack",
                 "20 reps total (10 per side)",
                 "Start at half-court. Dribble at controlled pace, hesitate — slow down, sit, drop your shoulders like you're about to settle — then explode past an imaginary defender. Right side first, then left."),
                ("Crossover at speed",
                 "8 trips down, one move per trip",
                 "Half-court to baseline at game speed. Each trip pick ONE move at the elbow area: crossover, between the legs, behind the back. Keep your eyes up the entire trip."),
            ],
        },
        {
            "day_number": 3,
            "day_title": "Two-Ball + Game Speed",
            "day_total_minutes": 15,
            "drills": [
                ("Two-ball synchronized pound",
                 "3 sets of 45 sec",
                 "Hold one ball in each hand. Pound dribble both at the same time at hip height. Both bounces hit the floor together. Builds even hand strength and rhythm."),
                ("Two-ball alternating",
                 "3 sets of 45 sec",
                 "One ball goes up while the other goes down. Alternating dribbles. Will feel awkward at first — that's the point. Eyes up forces you to feel the rhythm instead of looking."),
                ("1-on-0 attack from top",
                 "10 reps — 5 going right, 5 going left",
                 "Start at top of key. Pick ONE move (crossover, between legs, behind back, or hesi) and attack the rim at full speed. Finish strong at the rim. This is the game rep — make every one count."),
            ],
        },
    ],
}

# Which programs go in which sheet
PROGRAMS_BY_SHEET = {
    "1 - PG": [PG_SHOOTING, PG_HANDLE],
    "2 - SG": [],
    "3 - SF": [],
    "4 - PF": [],
    "5 - C":  [],
}


def write_header_rows(ws):
    """Add the orange header row + gray hint row to a sheet."""
    header_fill = PatternFill("solid", fgColor="FF7A1A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    hint_font = Font(italic=True, color="6B7280", size=9)
    for i, (name, width, hint) in enumerate(COLUMNS, start=1):
        # Row 1: header
        h_cell = ws.cell(row=1, column=i, value=name)
        h_cell.font = header_font
        h_cell.fill = header_fill
        h_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
        # Row 2: hint
        hint_cell = ws.cell(row=2, column=i, value=hint)
        hint_cell.font = hint_font
        hint_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"


def write_program(ws, start_row: int, program) -> int:
    """Write all drill rows for a program. Returns the next free row."""
    row = start_row
    for day in program["days"]:
        for j, (drill_name, drill_reps, drill_desc) in enumerate(day["drills"], start=1):
            row_data = {
                **program["meta"],
                "day_number": day["day_number"],
                "day_title": day["day_title"],
                "day_total_minutes": day["day_total_minutes"],
                "drill_order": j,
                "drill_name": drill_name,
                "drill_reps": drill_reps,
                "drill_description": drill_desc,
                "drill_video_url": "",
            }
            for col_idx, (col_name, _w, _h) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row, column=col_idx, value=row_data[col_name])
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
    return row


def main() -> None:
    wb = openpyxl.Workbook()
    # Remove the default sheet that gets auto-created
    default = wb.active
    wb.remove(default)

    for sheet_name, _position_label in SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        write_header_rows(ws)
        # Write programs assigned to this sheet
        next_row = 3
        for prog in PROGRAMS_BY_SHEET.get(sheet_name, []):
            next_row = write_program(ws, next_row, prog)
            next_row += 1   # blank row between programs

    wb.save(OUT)
    print(f"✓ wrote {OUT}")
    print(f"  5 sheets: {', '.join(s for s, _ in SHEETS)}")
    n_programs = sum(len(p) for p in PROGRAMS_BY_SHEET.values())
    print(f"  Seeded with {n_programs} example programs in '1 - PG' sheet")


if __name__ == "__main__":
    main()
